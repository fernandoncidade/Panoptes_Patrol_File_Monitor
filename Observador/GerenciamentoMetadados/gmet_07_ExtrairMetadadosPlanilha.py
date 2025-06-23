import os

def extrair_metadados_planilha(caminho, loc):
    metadados = {}
    ext = os.path.splitext(caminho)[1].lower()

    try:
        if ext in ['.xlsx', '.xlsm', '.xls']:
            try:
                if ext in ['.xlsx', '.xlsm']:
                    from openpyxl import load_workbook
                    wb = load_workbook(caminho, read_only=True, data_only=True)

                    planilhas = len(wb.sheetnames)
                    metadados['planilhas'] = planilhas
                    metadados['nomes_planilhas'] = ", ".join(wb.sheetnames)

                    linhas_total = 0
                    colunas_total = 0

                    for sheet_name in wb.sheetnames[:3]:
                        sheet = wb[sheet_name]
                        if hasattr(sheet, 'max_row') and hasattr(sheet, 'max_column'):
                            linhas_total += sheet.max_row
                            if sheet.max_column > colunas_total:
                                colunas_total = sheet.max_column

                    is_protected = False
                    for sheet_name in wb.sheetnames:
                        sheet = wb[sheet_name]
                        if hasattr(sheet, 'protection') and sheet.protection.sheet:
                            is_protected = True
                            break

                    if is_protected:
                        metadados['protegido'] = loc.get_text("yes") + " (planilhas protegidas)"

                    if wb.properties:
                        if wb.properties.creator:
                            metadados['autor'] = wb.properties.creator

                        if wb.properties.title:
                            metadados['titulo'] = wb.properties.title

                        if wb.properties.created:
                            metadados['data_criacao_doc'] = str(wb.properties.created)

                        if wb.properties.modified:
                            metadados['data_mod_doc'] = str(wb.properties.modified)

                    wb.close()

                elif ext == '.xls':
                    import xlrd
                    wb = xlrd.open_workbook(caminho, on_demand=True)
                    planilhas = len(wb.sheet_names())
                    metadados['planilhas'] = planilhas
                    metadados['nomes_planilhas'] = ", ".join(wb.sheet_names())

                    linhas_total = 0
                    colunas_total = 0

                    for idx in range(min(3, planilhas)):
                        sheet = wb.sheet_by_index(idx)
                        linhas_total += sheet.nrows
                        if sheet.ncols > colunas_total:
                            colunas_total = sheet.ncols

                    metadados['total_linhas'] = str(linhas_total)
                    metadados['colunas'] = str(colunas_total)

                    if hasattr(wb, 'protection_mode') and wb.protection_mode:
                        metadados['protegido'] = loc.get_text("yes")

                    wb.release_resources()

            except Exception as e:
                print(f"Erro ao extrair metadados da planilha {caminho}: {e}")

        elif ext == '.csv':
            try:
                import csv
                linhas = 0
                colunas = 0

                with open(caminho, 'r', encoding='utf-8', errors='ignore') as f:
                    reader = csv.reader(f)
                    primeira_linha = next(reader, None)
                    if primeira_linha:
                        colunas = len(primeira_linha)

                    linhas = 1
                    for _ in reader:
                        linhas += 1

                metadados['planilhas'] = 1
                nome_arquivo = os.path.basename(caminho)
                nome_sem_ext = os.path.splitext(nome_arquivo)[0]
                metadados['nomes_planilhas'] = nome_sem_ext

                metadados['total_linhas'] = linhas
                metadados['colunas'] = colunas

            except Exception as e:
                print(f"Erro ao extrair metadados do CSV {caminho}: {e}")

    except Exception as e:
        print(f"Erro geral ao extrair metadados da planilha {caminho}: {e}")

    return metadados
