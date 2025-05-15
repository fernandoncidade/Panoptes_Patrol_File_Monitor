import os
from .gmet_13_ExtrairMetadadosBackup import eh_arquivo_texto, contar_linhas
from .gmet_18_GetMetadados import get_metadados
from .gmet_19_GetTipoArquivo import identificar_tipo_arquivo
from .gmet_21_GetFormataTamanho import formata_tamanho

def get_dimensoes_arquivo(self, item, loc):
    caminho = item.get("dir_atual") or item.get("dir_anterior")
    if not caminho or not os.path.exists(caminho):
        if "dimensoes" in item and item["dimensoes"]:
            return loc.traduzir_metadados(item["dimensoes"], "dimensoes")
        return ""

    ext = os.path.splitext(caminho)[1].lower()
    tipo = identificar_tipo_arquivo(caminho, loc)

    if tipo == loc.get_text("file_image") or tipo == loc.get_text("file_video"):
        with self.lock_cache:
            if caminho in self.cache_metadados and 'dimensoes' in self.cache_metadados[caminho]:
                return self.cache_metadados[caminho]['dimensoes']

    with self.lock_cache:
        if caminho in self.cache_metadados and 'dimensoes' in self.cache_metadados[caminho]:
            return ""

    ext = os.path.splitext(caminho)[1].lower()
    nome_arquivo = os.path.basename(caminho).lower()

    try:
        if ext == '.bak' in nome_arquivo or 'backup' in nome_arquivo or 'bkp' in nome_arquivo:
            if eh_arquivo_texto(caminho):
                num_linhas = contar_linhas(caminho)
                with self.lock_cache:
                    if caminho not in self.cache_metadados:
                        self.cache_metadados[caminho] = {}

                    self.cache_metadados[caminho]["linhas"] = str(num_linhas)

                return ""

            else:
                tamanho = os.path.getsize(caminho)
                with self.lock_cache:
                    if caminho not in self.cache_metadados:
                        self.cache_metadados[caminho] = {}

                    self.cache_metadados[caminho]["binario"] = formata_tamanho(tamanho)

                return ""

        elif ext == '.log' or 'log' in nome_arquivo.lower():
            from .gmet_14_ExtrairMetadadosLog import extrair_metadados_log
            metadados_log = extrair_metadados_log(caminho, loc)

            if 'linhas' in metadados_log:
                with self.lock_cache:
                    if caminho not in self.cache_metadados:
                        self.cache_metadados[caminho] = {}

                    self.cache_metadados[caminho]["linhas"] = metadados_log["linhas"]

                return ""

        elif ext == '.dat':
            from .gmet_17_ExtrairMetadadosDadosEstruturados import extrair_metadados_dados_estruturados
            metadados_dat = extrair_metadados_dados_estruturados(caminho, loc)

            if 'registros' in metadados_dat and 'colunas' in metadados_dat:
                with self.lock_cache:
                    if caminho not in self.cache_metadados:
                        self.cache_metadados[caminho] = {}

                    self.cache_metadados[caminho]["registros"] = metadados_dat["registros"]
                    self.cache_metadados[caminho]["colunas"] = metadados_dat["colunas"]

                return ""

        elif ext == '.doc':
            import olefile
            if olefile.isOleFile(caminho):
                tamanho = os.path.getsize(caminho)
                paginas_estimadas = max(1, tamanho // 20000)

                with self.lock_cache:
                    if caminho not in self.cache_metadados:
                        self.cache_metadados[caminho] = {}

                    self.cache_metadados[caminho]["paginas_estimadas"] = str(paginas_estimadas)

                return ""

        elif ext == '.ppt':
            import olefile
            if olefile.isOleFile(caminho):
                tamanho = os.path.getsize(caminho)
                slides_estimados = max(1, tamanho // 100000)

                with self.lock_cache:
                    if caminho not in self.cache_metadados:
                        self.cache_metadados[caminho] = {}

                    self.cache_metadados[caminho]["slides_estimados"] = str(slides_estimados)

                return ""

        elif ext in ['.xlsx', '.xlsm']:
            from openpyxl import load_workbook
            wb = load_workbook(caminho, read_only=True, data_only=True)

            planilhas = len(wb.sheetnames)

            linhas_total = 0
            colunas_max = 0
            for sheet_name in wb.sheetnames[:3]:
                sheet = wb[sheet_name]
                if hasattr(sheet, 'max_row') and hasattr(sheet, 'max_column'):
                    linhas_total += sheet.max_row
                    if sheet.max_column > colunas_max:
                        colunas_max = sheet.max_column

            wb.close()

            with self.lock_cache:
                if caminho not in self.cache_metadados:
                    self.cache_metadados[caminho] = {}

                self.cache_metadados[caminho]["planilhas"] = str(planilhas)
                self.cache_metadados[caminho]["total_linhas"] = str(linhas_total)
                self.cache_metadados[caminho]["colunas"] = str(colunas_max)

            return ""

    except Exception as e:
        print(f"Erro ao extrair dimensões diretamente: {e}")

    metadados = get_metadados(item)

    if metadados:
        dimensoes_original = metadados.get("dimensoes", "")

        with self.lock_cache:
            if caminho not in self.cache_metadados:
                self.cache_metadados[caminho] = {}

            import re

            if re.search(r'(\d+)\s+(?:páginas|pages|páginas|pages|pagine|seiten)', dimensoes_original, re.IGNORECASE):
                match = re.search(r'(\d+)\s+(?:páginas|pages|páginas|pages|pagine|seiten)', dimensoes_original, re.IGNORECASE)
                self.cache_metadados[caminho]["paginas"] = match.group(1)

            if re.search(r'(\d+)\s+(?:páginas estimadas|estimated pages|páginas estimadas|pages estimées|pagine stimate|geschätzte seiten)', dimensoes_original, re.IGNORECASE):
                match = re.search(r'(\d+)\s+(?:páginas estimadas|estimated pages|páginas estimadas|pages estimées|pagine stimate|geschätzte seiten)', dimensoes_original, re.IGNORECASE)
                self.cache_metadados[caminho]["paginas_estimadas"] = match.group(1)

            if re.search(r'(\d+)\s+(?:linhas|lines|líneas|lignes|righe|zeilen)', dimensoes_original, re.IGNORECASE):
                match = re.search(r'(\d+)\s+(?:linhas|lines|líneas|lignes|righe|zeilen)', dimensoes_original, re.IGNORECASE)
                self.cache_metadados[caminho]["linhas"] = match.group(1)

            if re.search(r'(\d+)\s+(?:linhas de código|lines of code|líneas de código|lignes de code|righe di codice|codezeilen)', dimensoes_original, re.IGNORECASE):
                match = re.search(r'(\d+)\s+(?:linhas de código|lines of code|líneas de código|lignes de code|righe di codice|codezeilen)', dimensoes_original, re.IGNORECASE)
                self.cache_metadados[caminho]["linhas_codigo"] = match.group(1)

            if re.search(r'(\d+)\s+(?:total de linhas|total lines|total de líneas|total des lignes|totale righe|gesamtzeilen)', dimensoes_original, re.IGNORECASE):
                match = re.search(r'(\d+)\s+(?:total de linhas|total lines|total de líneas|total des lignes|totale righe|gesamtzeilen)', dimensoes_original, re.IGNORECASE)
                self.cache_metadados[caminho]["total_linhas"] = match.group(1)

            if re.search(r'(\d+)\s+(?:palavras|words|palabras|mots|parole|wörter)', dimensoes_original, re.IGNORECASE):
                match = re.search(r'(\d+)\s+(?:palavras|words|palabras|mots|parole|wörter)', dimensoes_original, re.IGNORECASE)
                self.cache_metadados[caminho]["palavras"] = match.group(1)

            if re.search(r'(\d+)\s+(?:slides|slides|diapositivas|diapositives|diapositive|folien)', dimensoes_original, re.IGNORECASE):
                match = re.search(r'(\d+)\s+(?:slides|slides|diapositivas|diapositives|diapositive|folien)', dimensoes_original, re.IGNORECASE)
                self.cache_metadados[caminho]["slides"] = match.group(1)

            if re.search(r'(\d+)\s+(?:slides estimados|estimated slides|diapositivas estimadas|diapositives estimées|diapositive stimate|geschätzte folien)', dimensoes_original, re.IGNORECASE):
                match = re.search(r'(\d+)\s+(?:slides estimados|estimated slides|diapositivas estimadas|diapositives estimées|diapositive stimate|geschätzte folien)', dimensoes_original, re.IGNORECASE)
                self.cache_metadados[caminho]["slides_estimados"] = match.group(1)

            if re.search(r'(\d+)\s+(?:planilhas|spreadsheets|hojas de cálculo|feuilles de calcul|fogli di calcolo|tabellenkalkulationen)', dimensoes_original, re.IGNORECASE):
                match = re.search(r'(\d+)\s+(?:planilhas|spreadsheets|hojas de cálculo|feuilles de calcul|fogli di calcolo|tabellenkalkulationen)', dimensoes_original, re.IGNORECASE)
                self.cache_metadados[caminho]["planilhas"] = match.group(1)

            if re.search(r'(\d+)\s+(?:colunas|columns|columnas|colonnes|colonne|spalten)', dimensoes_original, re.IGNORECASE):
                match = re.search(r'(\d+)\s+(?:colunas|columns|columnas|colonnes|colonne|spalten)', dimensoes_original, re.IGNORECASE)
                self.cache_metadados[caminho]["colunas"] = match.group(1)

            if re.search(r'(\d+)\s+(?:arquivos|files|archivos|fichiers|file|dateien)', dimensoes_original, re.IGNORECASE):
                match = re.search(r'(\d+)\s+(?:arquivos|files|archivos|fichiers|file|dateien)', dimensoes_original, re.IGNORECASE)
                self.cache_metadados[caminho]["arquivos"] = match.group(1)

            if re.search(r'([\d.,]+\s+[KMGT]?B)\s+(?:descompactado|unzipped|descomprimido|décompressé|decompresso|entpackt)', dimensoes_original, re.IGNORECASE):
                match = re.search(r'([\d.,]+\s+[KMGT]?B)\s+(?:descompactado|unzipped|descomprimido|décompressé|decompresso|entpackt)', dimensoes_original, re.IGNORECASE)
                self.cache_metadados[caminho]["descompactados"] = match.group(1)

            if re.search(r'(\d+)\s+(?:tabelas|tables|tablas|tables|tabelle|tabellen)', dimensoes_original, re.IGNORECASE):
                match = re.search(r'(\d+)\s+(?:tabelas|tables|tablas|tables|tabelle|tabellen)', dimensoes_original, re.IGNORECASE)
                self.cache_metadados[caminho]["tabelas"] = match.group(1)

            if re.search(r'(\d+)\s+(?:registros|records|registros|enregistrements|record|datensätze)', dimensoes_original, re.IGNORECASE):
                match = re.search(r'(\d+)\s+(?:registros|records|registros|enregistrements|record|datensätze)', dimensoes_original, re.IGNORECASE)
                self.cache_metadados[caminho]["registros"] = match.group(1)

            binario_match = re.search(r'(?:binário|binary|binario|fichier binaire|binario|binär):\s+([\d.,]+\s+[KMGT]?B)', dimensoes_original, re.IGNORECASE)
            if binario_match:
                self.cache_metadados[caminho]["binario"] = binario_match.group(1)

    return ""
