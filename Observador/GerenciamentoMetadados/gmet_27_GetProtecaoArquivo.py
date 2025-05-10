import os
import win32file
import win32con

def get_protecao_arquivo(gerenciador, item, loc):
    caminho = item.get("dir_atual") or item.get("dir_anterior")
    if not caminho or not os.path.exists(caminho):
        if "protegido" in item and item["protegido"]:
            return loc.traduzir_metadados(item["protegido"], "protegido")
        return ""

    try:
        with gerenciador.lock_cache:
            if caminho in gerenciador.cache_metadados and "protegido" in gerenciador.cache_metadados[caminho]:
                protegido = gerenciador.cache_metadados[caminho]["protegido"]
                return loc.traduzir_metadados(protegido, "protegido")

        attrs = win32file.GetFileAttributes(caminho)
        is_readonly = bool(attrs & win32con.FILE_ATTRIBUTE_READONLY)
        is_hidden = bool(attrs & win32con.FILE_ATTRIBUTE_HIDDEN)
        is_system = bool(attrs & win32con.FILE_ATTRIBUTE_SYSTEM)
        is_encrypted = bool(attrs & win32con.FILE_ATTRIBUTE_ENCRYPTED)
        is_compressed = bool(attrs & win32con.FILE_ATTRIBUTE_COMPRESSED)

        protecao = []
        if is_readonly or is_system or is_encrypted or is_hidden:
            protecao.append(loc.get_text("yes"))

            if is_readonly:
                protecao.append(loc.get_text("readonly"))

            if is_hidden:
                protecao.append(loc.get_text("hidden"))

            if is_system:
                protecao.append(loc.get_text("system"))

            if is_encrypted:
                protecao.append(loc.get_text("encrypted"))

            if is_compressed:
                protecao.append(loc.get_text("compressed"))

            ext = os.path.splitext(caminho)[1].lower()
            if ext in ['.docx', '.xlsx', '.pptx', '.doc', '.xls', '.ppt']:
                try:
                    if ext == '.docx':
                        from docx import Document
                        try:
                            Document(caminho)

                        except Exception as e:
                            if "document is encrypted" in str(e).lower():
                                protecao.append(loc.get_text("password_protected"))

                    elif ext == '.xlsx':
                        from openpyxl import load_workbook
                        try:
                            wb = load_workbook(caminho, read_only=True, data_only=True)
                            for sheet_name in wb.sheetnames:
                                sheet = wb[sheet_name]
                                if hasattr(sheet, 'protection') and sheet.protection.sheet:
                                    protecao.append(loc.get_text("sheet_protected"))
                                    break

                            wb.close()

                        except Exception as e:
                            if "file is encrypted" in str(e).lower():
                                protecao.append(loc.get_text("password_protected"))

                    elif ext == '.pptx':
                        from pptx import Presentation
                        try:
                            prs = Presentation(caminho)

                        except Exception as e:
                            if "is encrypted" in str(e).lower():
                                protecao.append(loc.get_text("password_protected"))

                    elif ext in ['.doc', '.xls', '.ppt']:
                        import olefile
                        if olefile.isOleFile(caminho):
                            try:
                                with olefile.OleFile(caminho) as ole:
                                    pass

                            except Exception as e:
                                if "encrypted" in str(e).lower():
                                    protecao.append(loc.get_text("password_protected"))

                except ImportError:
                    pass

            return ", ".join(protecao)

        else:
            ext = os.path.splitext(caminho)[1].lower()
            if ext in ['.docx', '.xlsx', '.pptx', '.doc', '.xls', '.ppt']:
                try:
                    if ext == '.docx':
                        from docx import Document
                        try:
                            Document(caminho)

                        except Exception as e:
                            if "document is encrypted" in str(e).lower():
                                protecao.append(loc.get_text("password_protected"))

                    elif ext == '.xlsx':
                        from openpyxl import load_workbook
                        try:
                            wb = load_workbook(caminho, read_only=True, data_only=True)
                            for sheet_name in wb.sheetnames:
                                sheet = wb[sheet_name]
                                if hasattr(sheet, 'protection') and sheet.protection.sheet:
                                    protecao.append(loc.get_text("sheet_protected"))
                                    break

                            wb.close()

                        except Exception as e:
                            if "file is encrypted" in str(e).lower():
                                protecao.append(loc.get_text("password_protected"))

                    elif ext == '.pptx':
                        from pptx import Presentation
                        try:
                            prs = Presentation(caminho)

                        except Exception as e:
                            if "is encrypted" in str(e).lower():
                                protecao.append(loc.get_text("password_protected"))

                    elif ext in ['.doc', '.xls', '.ppt']:
                        import olefile
                        if olefile.isOleFile(caminho):
                            try:
                                with olefile.OleFile(caminho) as ole:
                                    pass

                            except Exception as e:
                                if "encrypted" in str(e).lower():
                                    protecao.append(loc.get_text("password_protected"))

                except ImportError:
                    pass

            return loc.get_text("no")

    except Exception as e:
        if "protegido" in item and item["protegido"]:
            return loc.traduzir_metadados(item["protegido"], "protegido")

        print(f"Erro ao verificar proteção: {e}")
        return ""
